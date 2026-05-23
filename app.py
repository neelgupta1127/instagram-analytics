# =========================================================
# INSTAGRAM FOLLOWERS / FOLLOWING ANALYTICS TOOL
# Streamlit Web App — Fixed Version
# =========================================================

import streamlit as st
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="Instagram Analytics",
    page_icon="📊",
    layout="centered"
)

# =========================================================
# CUSTOM STYLE
# =========================================================

st.markdown("""
    <style>
        .stDownloadButton button {
            background-color: #E1306C;
            color: white;
            font-weight: bold;
            border-radius: 8px;
            width: 100%;
        }
        .stMetric {
            background-color: #1a1a2e;
            border-radius: 10px;
            padding: 10px;
        }
    </style>
""", unsafe_allow_html=True)

# =========================================================
# HEADER
# =========================================================

st.title("📊 Instagram Followers Analytics")
st.markdown("Upload your **Instagram exported HTML files** to analyze your followers and following.")
st.markdown("---")

# =========================================================
# FIXED USERNAME EXTRACTION FUNCTION
# Handles both plain usernames AND full Instagram URLs
# =========================================================

def extract_usernames(html_bytes):
    """
    Extracts Instagram usernames from exported HTML files.
    - followers_1.html stores plain usernames as link text
    - following.html stores full URLs as link text
    This function handles BOTH formats correctly.
    """
    soup = BeautifulSoup(html_bytes, 'lxml')
    usernames = set()

    for link in soup.find_all('a'):
        text = link.get_text(strip=True)
        href = link.get('href', '')

        if text.startswith('http'):
            # following.html format — extract username from URL path
            username = href.rstrip('/').split('/')[-1]
            if not username:
                # fallback: extract from text URL
                username = text.rstrip('/').split('/')[-1]
            if username:
                usernames.add(username.lower())
        elif text:
            # followers_1.html format — plain username text
            usernames.add(text.lower())

    return usernames

# =========================================================
# FILE UPLOAD
# =========================================================

st.subheader("📁 Upload Instagram Export Files")
st.info("Go to Instagram → Settings → Account → Download your data → Request ZIP → Extract and upload the two HTML files below.")

col1, col2 = st.columns(2)
with col1:
    followers_file = st.file_uploader("📥 Upload followers_1.html", type=["html", "htm"])
with col2:
    following_file = st.file_uploader("📤 Upload following.html", type=["html", "htm"])

# =========================================================
# ANALYSIS
# =========================================================

if followers_file and following_file:

    with st.spinner("🔍 Extracting usernames and running analysis..."):

        followers_set  = extract_usernames(followers_file.read())
        following_set  = extract_usernames(following_file.read())

        mutual             = sorted(followers_set & following_set)
        not_following_back = sorted(following_set - followers_set)
        fans               = sorted(followers_set - following_set)

    st.success("✅ Analysis Complete!")
    st.markdown("---")

    # =========================================================
    # SUMMARY METRICS
    # =========================================================

    st.subheader("📈 Summary")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("👥 Followers",          len(followers_set))
    c2.metric("➡️ Following",           len(following_set))
    c3.metric("🤝 Mutual",             len(mutual))
    c4.metric("❌ Not Following Back",  len(not_following_back))
    c5.metric("⭐ Fans",               len(fans))

    st.markdown("---")

    # =========================================================
    # SET THEORY REFERENCE
    # =========================================================

    with st.expander("📐 How the analysis works (Set Theory)"):
        st.markdown("""
        | Category | Formula | Meaning |
        |---|---|---|
        | 🤝 Mutual Followers | A ∩ B | You follow them AND they follow you |
        | ❌ Not Following Back | B - A | You follow them but they DON'T follow you |
        | ⭐ Fans | A - B | They follow you but you DON'T follow them |

        **A = Followers Set &nbsp;&nbsp; B = Following Set**
        """)

    # =========================================================
    # RESULTS TABS
    # =========================================================

    st.subheader("📋 Detailed Results")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        f"🤝 Mutual ({len(mutual)})",
        f"❌ Not Following Back ({len(not_following_back)})",
        f"⭐ Fans ({len(fans)})",
        f"👥 All Followers ({len(followers_set)})",
        f"➡️ All Following ({len(following_set)})"
    ])

    with tab1:
        st.caption("Both you and these accounts follow each other.")
        st.dataframe(
            pd.DataFrame(mutual, columns=["Username"]),
            use_container_width=True, hide_index=True
        )

    with tab2:
        st.caption("You follow these accounts but they do NOT follow you back.")
        st.dataframe(
            pd.DataFrame(not_following_back, columns=["Username"]),
            use_container_width=True, hide_index=True
        )

    with tab3:
        st.caption("These accounts follow you but you do NOT follow them back.")
        st.dataframe(
            pd.DataFrame(fans, columns=["Username"]),
            use_container_width=True, hide_index=True
        )

    with tab4:
        st.dataframe(
            pd.DataFrame(sorted(followers_set), columns=["Username"]),
            use_container_width=True, hide_index=True
        )

    with tab5:
        st.dataframe(
            pd.DataFrame(sorted(following_set), columns=["Username"]),
            use_container_width=True, hide_index=True
        )

    st.markdown("---")

    # =========================================================
    # GENERATE & DOWNLOAD EXCEL REPORT
    # =========================================================

    st.subheader("📥 Download Full Excel Report")

    output = io.BytesIO()

    summary_df = pd.DataFrame({
        "Metric": [
            "Total Followers",
            "Total Following",
            "Mutual Followers",
            "Not Following Back",
            "Fans"
        ],
        "Count": [
            len(followers_set),
            len(following_set),
            len(mutual),
            len(not_following_back),
            len(fans)
        ]
    })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer,
            sheet_name='Summary', index=False)
        pd.DataFrame(mutual, columns=["Mutual Followers"]).to_excel(writer,
            sheet_name='Mutual Followers', index=False)
        pd.DataFrame(not_following_back, columns=["Not Following Back"]).to_excel(writer,
            sheet_name='Not Following Back', index=False)
        pd.DataFrame(fans, columns=["Fans"]).to_excel(writer,
            sheet_name='Fans', index=False)
        pd.DataFrame(sorted(followers_set), columns=["Followers"]).to_excel(writer,
            sheet_name='All Followers', index=False)
        pd.DataFrame(sorted(following_set), columns=["Following"]).to_excel(writer,
            sheet_name='All Following', index=False)

    # Format Excel — bold headers, color, auto column width
    output.seek(0)
    workbook = load_workbook(output)

    header_fill = PatternFill(start_color="E1306C", end_color="E1306C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 6

    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    st.download_button(
        label="📥 Download Instagram_Analytics_Report.xlsx",
        data=final_output,
        file_name="Instagram_Analytics_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.warning("👆 Please upload BOTH files above to start the analysis.")

# =========================================================
# FOOTER
# =========================================================

st.markdown("---")
st.markdown(
    "<center>🔒 100% Offline · No login required · No scraping · Official Instagram export data only</center>",
    unsafe_allow_html=True
)
